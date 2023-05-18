# -*- coding: utf8 -*-
import os
import pandas as pd
from shutil import copy2
import openpyxl
from openpyxl.styles import Alignment, Border, Side, PatternFill, Font
from openpyxl.formatting.formatting import ConditionalFormattingList
import re


def filling_in_excel(df, wb, first_row=1, first_col=1, ws_title=None, color_pattern=None, extra=None):
    '''
    Функция заполняет шаблонный лист excel данными из датафрейма.
    В качестве параметра (color_pattern) можно передать словарь вида {'паттерн': 'цвет hex'} для 
    закрашивания ячеек выбранным цветом, если паттерн подходит.
    Параметры:
    df - датафрэйм;
    wb - WorkBook-объект (не имя файла!)
    border - параметры границ ячеек
    out_path -  путь к результирующему файлу (по умолчанию ='result.xlsx');
    first_row - начальная строка заполнения, начиная с заголовка (по умолчанию =1);
    first_col - начальная колонка заполнения (по умолчанию =1);
    color_pattern - словарь вида {'паттерн': 'цвет hex'} для закрашивания ячеек цветом.
    extra - словарь вида {'ячейка': 'значение'} для вставки дополнительной информации. Ячейка - в формате 'A1' и т.п.
    '''
    #wb = openpyxl.load_workbook(template_path)                                   # открываем файл excel
    if ws_title == None:                                                         # указываем название вкладки, куда будем записывать
        ws = wb.active
    else:
        ws = wb[ws_title]
    
    thin = Side(border_style="thin", color="000000")                             # стиль 'thin': линии тонкие, цвет чёрный
    border = Border(top=thin, left=thin, right=thin, bottom=thin)                # параметры границ ячеек

    i = first_row                                                                # начальная строка заполнения
    
    for header_no in range(0, len(df.columns)):
        ws_cell = ws.cell(row= i, column = header_no + 1)
        ws_cell.value = df.columns[header_no]
        ws_cell.border = border
        ws_cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
        
    i += 1                                                                       # вторая строка после заголовка имеет номер +1
    for idx, row in df.iterrows():                                               # итерируемся по строкам
        j = first_col                                                            # начальная колонка заполнения
        for val in row:                                                          # итерируемся по всем значениям кортежа row
            ws_cell = ws.cell(row= i, column = j)                                # выбираем ячейку по номеру строки и колонки
            ws_cell.value = val                                                  # присваиваем значение ячейке
            ws_cell.border = border                                              # устанавливаем границы ячейки
            if color_pattern != None:                                            # если параметр color_pattern указан, закрашиваем цветом
                for pattern, fill_color in color_pattern.items():                # итерируемся по словарю, извлекая паттерны и цвета
                    if re.search(pattern, str(ws_cell.value)) != None:           # если нашлись совпаденяия (т.е. != None), красим в цвет
                        ws_cell.fill = PatternFill("solid", fgColor=fill_color)  # параметры заполнения цветом
            j += 1
        i += 1
    if extra != None:                                                            # итеритуемся по ключам и значениям словаря extra
        for key, value in extra.items():
            ws[key] = value                                                      # записываем значения в ячейку
    #wb.save(out_path)                                                            # сохраняем результат
    #wb.close() 

#ws.conditional_formatting = ConditionalFormattingList()                      # очищаем условное форматирование (окрашивание цветом ячеек в зависимости от значений)

files = os.listdir('.')   # получаем список всех файлов в директории 

# ищем excel файл
for x in files:
    if x[-5:].lower() == '.xlsx':
        xl = x

# считываем excel
df = pd.read_excel(xl, skiprows=2)

# ------

# удаляем лишние колонки
df.drop(columns=['Кол-во секций', 'Кол-во на 1 секцию'], 
        inplace=True)

# удаляем строки с пропусками в колонке 'Общее кол-во'
df.dropna(
     subset=['Общее кол-во'], 
     inplace=True
 )

# удаляем строки с нулевыми значениями в колонке 'Общее кол-во'
df = df[df['Общее кол-во'] != 0]

# сброс нумерации
df.reset_index(drop=True, inplace=True)  # сброс индекса
df['№'] = df.index + 1                   # колонка '№' равна индексу + 1

# проверяем дубликаты
duplix_df = (df
      .loc[df.iloc[:, 1]
           .duplicated()]
      )

# печатаем информацию о повторяющихся наименованиях
if len(duplix_df) == 0:
    print('Дубликатов нет')
else:
    print('Дубликаты:')
    #print(duplix_df
          #.iloc[:, 1]
          #.to_string(index=False))
    for i, row in duplix_df.iterrows():
        print(row.iloc[1])
print()

# создаём папку для dxf
new_dxf_folder = 'DXF для заказа'       # папка для отобранных файлов DXF
if not os.path.exists(new_dxf_folder):
    os.makedirs(new_dxf_folder)


# копируем dxf
# итеритуемся по второй колонке датафрэйма
for dxf in df.iloc[:,1]:        
    try:
        copy2(dxf + '.dxf', new_dxf_folder +'/' + dxf + '.dxf')         
    except:
        try:
            copy2(dxf + '.DXF', new_dxf_folder +'/' + dxf + '.dxf')
        except:
            print('Нет файла:', dxf)


excel_path = new_dxf_folder +'/' + 'final.xlsx'    # путь к финальному файлу Excel

wb = openpyxl.Workbook()                   # создаём пустой файл Excel
ws = wb.active                             # лист
ws.column_dimensions['A'].width = 4.67     # устанавливаем ширину ячеек (в ещиницах шрифта)
ws.column_dimensions['B'].width = 47.67
ws.column_dimensions['C'].width = 14.22
ws.column_dimensions['D'].width = 9.22
ws.column_dimensions['E'].width = 8.33
ws.column_dimensions['F'].width = 6.89
ws.column_dimensions['G'].width = 8.11

ws.cell(row= 1, column = 5).value = 'Бокин'       # присваиваем значение ячейкам
ws.cell(row= 2, column = 2).value = 'Заявка №'
ws.cell(row= 2, column = 2).font = Font(size=18)  # устанавливаем размер шрифта ячейки


filling_in_excel(df, wb, first_row=3, first_col=1) # заполняем Excel данными из df с помощью функции


wb.save(excel_path)                                                            # сохраняем результат
wb.close()



print()
input('Нажмите Enter чтобы выйти')


